"""create_dummy_answers.py

Generate a synthetic TWF2026_回答集約.xlsx for local testing of excel_mapper.py
without needing the real (sensitive) maker reply data.

5 makers cover the matching cases we want to exercise:
  66 ㈱ダイヘン           - normal exact-after-normalize match
  32 興研㈱                - normal match (legal-form swap)
  47 ㈱重松製作所         - normal match
  86 日鉄溶接工業㈱       - normalization needed (CSV uses 康熙部首 ⽇ U+2F00,
                            Excel uses normal 日 U+65E5)
  58 3Mジャパン            - intentional UNMATCH (CSV name is スリーエムジャパン㈱)
                            -> exercises the unmatched-warning path
"""
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / "data" / "raw" / "answers_dummy.xlsx"
OUT.parent.mkdir(parents=True, exist_ok=True)

HEADERS = [
    "No.", "回答状況", "メーカー名", "宛名", "メールアドレス",
    "送信区分", "返信者メールアドレス", "受信日時",
    "回答1", "回答2", "回答3", "回答4", "回答5",
    "添付数", "添付ファイル名", "添付フォルダパス", "備考",
]

ROWS = [
    [
        66, "☑", "㈱ダイヘン", "営業企画部", "daihen@example.co.jp", "個別",
        "tanaka@daihen.co.jp", datetime(2026, 4, 25, 14, 32),
        "新型溶接ロボット FD-V8 を中央ステージで実演。30分毎の定時実演 + 個別商談ブース併設。",
        "AIアーク制御を搭載した新型 TIG溶接機 / 高速ファイバーレーザ溶接機 (新製品)。",
        "ロボット実演ステージ + デモ動画放映 / 来場者プレゼント (オリジナルUSBメモリ)。",
        "成約特典: 期間限定割引 + 5年保証延長プラン無償付帯。",
        "提案資料・カタログ一式を当日配布。商談予約推奨 (フォームリンク別途)。",
        2,
        "新型ロボット企画書.pdf\nFD-V8カタログ_2026.pdf",
        r"\\fileserver\twf2026\attachments\㈱ダイヘン",
        "回答テンプレ最新版で受領済",
    ],
    [
        32, "◎", "興研㈱", "営業推進部", "koken@example.co.jp", "個別",
        "saito@koken.co.jp", datetime(2026, 4, 22, 10, 15),
        "防塵マスク フィッティング体験ブース 両日終日開設。",
        "電動ファン付き呼吸用保護具 (PAPR) 新モデルを初公開。",
        "プロのフィッターによるマスク密着度測定サービス (1名10分)。",
        "成約30万円以上で予備フィルタ1年分プレゼント。",
        "粉塵作業現場での導入事例集を当日配布。",
        1,
        "PAPR新モデル仕様書.pdf",
        r"\\fileserver\twf2026\attachments\興研㈱",
        "",
    ],
    [
        47, "▲", "㈱重松製作所", "マーケティング部", "shigematsu@example.co.jp", "ユアサ一括",
        "yamada@shigematsu.co.jp", datetime(2026, 4, 28, 9, 0),
        "全面形面体マスク GM-26 試着ブース。",
        "電動ファン付き呼吸用保護具 サカヰ式。",
        "酸欠作業向け SCBA (自給式呼吸器) 体験コーナー。",
        "特になし",
        "事前来場予約優先。",
        0, "",
        r"\\fileserver\twf2026\attachments\㈱重松製作所",
        "回答薄い (▲) — 柏原追記が必要",
    ],
    [
        86, "☑", "日鉄溶接工業㈱", "技術営業課", "nittetsu@example.co.jp", "産メカ一括",
        "ito@nittetsu-yousetsu.co.jp", datetime(2026, 4, 30, 16, 45),
        "高張力鋼用フラックス入りワイヤ 最新ラインナップ。",
        "ステンレス鋼用 共金系ソリッドワイヤ 新製品。",
        "アーク特性比較デモ + サンプルワイヤ配布。",
        "ロット成約特典: 専用テクニカルサポート無償提供。",
        "造船・橋梁向け技術資料を当日配布。",
        3,
        "高張力鋼FCW_2026.pdf\nステンレス共金系SW.pdf\n技術資料一式.zip",
        r"\\fileserver\twf2026\attachments\日鉄溶接工業㈱",
        "CSVは康熙部首⽇ (U+2F00) 表記、Excelは通常の日 (U+65E5) — NFKC正規化で突合確認用",
    ],
    [
        # わざと CSV 表記 (スリーエムジャパン㈱) と乖離させた表記でテスト
        58, "☑", "3Mジャパン", "産業安全本部", "3m@example.co.jp", "個別",
        "suzuki@mmm.com", datetime(2026, 4, 21, 11, 30),
        "研磨ベルト・ディスク 新ライン製品。",
        "個人用保護具 (PPE) ヘルメット・保護メガネ・防塵マスクの提案。",
        "新型研磨機の体験ブース + 試し研磨。",
        "業界別ソリューション提案資料を配布。",
        "オンライン商談予約特典あり。",
        1,
        "3M_PPE_2026.pdf",
        r"\\fileserver\twf2026\attachments\スリーエムジャパン㈱",
        "Excelの社名 (3Mジャパン) は CSV (スリーエムジャパン㈱) と乖離 — 警告ログ確認用",
    ],
]


def main():
    wb = Workbook()
    ws = wb.active
    ws.title = "回答集約"
    ws.append(HEADERS)
    for r in ROWS:
        ws.append(r)

    widths = [6, 8, 22, 14, 24, 12, 26, 18, 40, 40, 40, 40, 40, 6, 30, 40, 36]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    wb.save(OUT)
    print(f"wrote {OUT.relative_to(ROOT)} ({OUT.stat().st_size:,} bytes)  rows={len(ROWS)}")


if __name__ == "__main__":
    main()
