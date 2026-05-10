"""excel_mapper.py

Read TWF2026_回答集約.xlsx (output by twf2026_collector / twf2026_sender) and
merge maker answers into the portal data layer:

  - data/maker_details.json  (148-maker structured data, keyed by zero-padded No.)
  - data/makers.csv          (update has_answer column; original backed up to .bak)

Usage:
  python scripts/excel_mapper.py [--excel PATH] [--dry]

Matching strategy:
  Excel No (送信順) と makers.csv No (TWFパンフ記載順) は番号体系が違うため
  No列ベースの突合は使わない。突合優先順位:
    1) メーカー名 normalize 完全一致 (NFKC + 法人格除去 + 空白除去 + casefold)
    2) maker_aliases.json による手動エイリアス
  突合失敗行は「Unmatched dump」に reply_email/email も併記して出力する
  (alias.json への追加候補や、makers.csv 外メーカーの判別に使う)。

  STATUS マーカーは sender 側 (commit e603089〜) の改修に追従:
    ☑ 回答済み / ◎ 部分回答 / 📎 添付回答 / ★ 複数返信 → has_answer=True
    ▲ 未記入返信 / ✖ 未回答                            → has_answer=False
    ■ 一括 (ユアサ/産メカ自体の dummy行) / ─ 経由      → skipped
"""
from __future__ import annotations

import argparse
import csv
import json
import re
import shutil
import sys
import unicodedata
from collections import Counter
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
CSV_PATH = ROOT / "data" / "makers.csv"
JSON_PATH = ROOT / "data" / "maker_details.json"
ALIAS_PATH = ROOT / "data" / "maker_aliases.json"
OVERRIDES_PATH = ROOT / "data" / "maker_overrides.json"
DEFAULT_EXCEL = Path(r"D:\repos\twf2026_sender\TWF2026_回答集約_local検証.xlsx")

# Excel layout (1-indexed columns, see twf2026_collector spec)
COL = {
    "no": 1, "status": 2, "name": 3,
    "addressee": 4, "email": 5, "send_type": 6,
    "reply_email": 7, "reply_dt": 8,
    "q1": 9, "q2": 10, "q3": 11, "q4": 12, "q5": 13,
    "attach_count": 14, "attach_files": 15, "attach_dir": 16, "note": 17,
}

LEGAL_RE = re.compile(
    r"(株式会社|有限会社|合同会社|合資会社|合名会社"
    r"|\(株\)|\(有\)|（株）|（有）"
    r"|㈱|㈲|㈳|㈴|㈵|㋿)"
)
WS_RE = re.compile(r"\s+")

# 添付ファイル: 残す拡張子 (案内資料系)
ATTACH_KEEP_EXT = {".pdf", ".docx", ".xlsx", ".pptx"}
# 添付ファイル: 明示的に除外するパターン (Outlook 埋め込み画像 / ロックファイル / 圧縮)
ATTACH_DROP_NAME_RE = re.compile(r"^image\d{3,}\.(gif|png|jpe?g|bmp)$", re.IGNORECASE)


def is_useful_attachment(filename: str) -> bool:
    """配布資料として意味ある添付だけ True を返す。
    - image001.gif / image123.png のような Outlook 埋め込み画像は除外
    - .txt / .zip は除外
    - .pdf / .docx / .xlsx / .pptx のみ許可
    """
    name = filename.strip()
    if not name:
        return False
    base = name.rsplit("/", 1)[-1].rsplit("\\", 1)[-1]
    if ATTACH_DROP_NAME_RE.match(base):
        return False
    ext = ("." + base.rsplit(".", 1)[-1].lower()) if "." in base else ""
    return ext in ATTACH_KEEP_EXT

# B列(回答状況)の先頭マーカー → 内部 status
# sender 側 (twf2026_collector commit e603089〜) の新ステータス体系に準拠。
STATUS_MAP = {
    "☑": "answered",     # 回答済
    "◎": "partial",      # 部分回答 (Q5/添付に意味ある内容あり)
    "📎": "partial",     # 添付回答 (本文空+添付のみ — 添付が回答の代わり)
    "★": "answered",     # 複数返信 (複数メールから本文を合体済)
    "▲": "unanswered",   # 未記入返信 (返信は来たが本文空)
    "✖": "unanswered",   # 未回答
    "■": "skipped",      # 一括ダミー (ユアサ/産メカ自体、集約対象外)
    "─": "skipped",      # 経由 (一括の派生記録、has_answer は本体行で記録)
}
HAS_ANSWER_STATUSES = {"answered", "partial"}


def classify_status(status: str) -> str:
    s = (status or "").strip()
    if not s:
        return "unknown"
    return STATUS_MAP.get(s[:1], "unknown")


def normalize(name) -> str:
    if name is None:
        return ""
    s = unicodedata.normalize("NFKC", str(name))
    s = LEGAL_RE.sub("", s)
    s = WS_RE.sub("", s)
    return s.casefold()


def cell_str(v) -> str:
    if v is None:
        return ""
    if isinstance(v, str):
        return v.strip()
    return str(v).strip()


def fmt_dt(v) -> str | None:
    if v is None or v == "":
        return None
    if isinstance(v, datetime):
        return v.strftime("%Y-%m-%d %H:%M")
    return str(v).strip()


def load_makers() -> list[dict]:
    with open(CSV_PATH, encoding="utf-8", newline="") as f:
        return list(csv.DictReader(f))


def load_overrides() -> dict[str, dict]:
    """data/maker_overrides.json を読んで {zero-padded No: {field: value, ...}} を返す。
    キーが '_' で始まるエントリ・各 entry 内の '_' で始まるフィールドはコメント扱いで無視する。
    """
    if not OVERRIDES_PATH.exists():
        return {}
    with open(OVERRIDES_PATH, encoding="utf-8") as f:
        raw = json.load(f)
    out: dict[str, dict] = {}
    for k, v in raw.items():
        if k.startswith("_"):
            continue
        if not isinstance(v, dict):
            continue
        out[k] = {fk: fv for fk, fv in v.items() if not fk.startswith("_")}
    return out


def load_aliases() -> dict[str, str]:
    """data/maker_aliases.json を読んで {normalize(excel_name): normalize(csv_name)} を返す。
    キーが '_' で始まるエントリはドキュメント扱いで無視する。
    """
    if not ALIAS_PATH.exists():
        return {}
    with open(ALIAS_PATH, encoding="utf-8") as f:
        raw = json.load(f)
    return {
        normalize(k): normalize(v)
        for k, v in raw.items()
        if not k.startswith("_")
    }


def load_excel(path: Path) -> list[dict]:
    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb["回答集約"] if "回答集約" in wb.sheetnames else wb.active
    rows: list[dict] = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue  # header
        if not row:
            continue
        no_cell = row[COL["no"] - 1]
        if no_cell in (None, ""):
            continue
        try:
            no = int(no_cell)
        except (TypeError, ValueError):
            continue
        rec = {
            "no": no,
            "status": cell_str(row[COL["status"] - 1]),
            "name": cell_str(row[COL["name"] - 1]),
            "email": cell_str(row[COL["email"] - 1]),
            "reply_email": cell_str(row[COL["reply_email"] - 1]),
            "reply_dt": row[COL["reply_dt"] - 1],
            "q1": cell_str(row[COL["q1"] - 1]),
            "q2": cell_str(row[COL["q2"] - 1]),
            "q3": cell_str(row[COL["q3"] - 1]),
            "q4": cell_str(row[COL["q4"] - 1]),
            "q5": cell_str(row[COL["q5"] - 1]),
            "attach_files_raw": cell_str(row[COL["attach_files"] - 1]),
            "attach_dir": cell_str(row[COL["attach_dir"] - 1]),
            "note": cell_str(row[COL["note"] - 1]),
        }
        rows.append(rec)
    wb.close()
    return rows


def init_details(makers: list[dict]) -> dict:
    out = {}
    for m in makers:
        no = int(m["no"])
        out[f"{no:03d}"] = {
            "no": no,
            "name": m["name"],
            "name_short": m.get("name_short", ""),
            "category": m.get("category", ""),
            "has_answer": False,
            "status": "unknown",
            "reply_date": None,
            "q1": "", "q2": "", "q3": "", "q4": "", "q5": "",
            "attachments": [],
            "attachment_dir": None,
            "company_dir": None,
        }
    return out


def merge(makers: list[dict], excel_rows: list[dict]):
    by_norm = {normalize(m["name"]): m for m in makers}
    aliases = load_aliases()
    details = init_details(makers)
    matched: list[dict] = []
    unmatched: list[dict] = []
    raw_status_counter: Counter = Counter()
    classified_counter: Counter = Counter()
    aliased: list[dict] = []

    for r in excel_rows:
        raw_status = r.get("status", "")
        classified = classify_status(raw_status)
        raw_status_counter[raw_status] += 1
        classified_counter[classified] += 1

        norm_name = normalize(r["name"])
        match = by_norm.get(norm_name)
        if not match:
            target_norm = aliases.get(norm_name)
            if target_norm:
                match = by_norm.get(target_norm)
                if match:
                    aliased.append({
                        "excel_no": r["no"], "excel_name": r["name"],
                        "csv_no": int(match["no"]), "csv_name": match["name"],
                        "status": raw_status, "classified": classified,
                    })
        if not match:
            unmatched.append({
                "excel_no": r["no"],
                "excel_name": r["name"],
                "email": r.get("email") or None,
                "reply_email": r.get("reply_email") or None,
                "status": raw_status,
                "classified": classified,
            })
            continue

        no = int(match["no"])
        key = f"{no:03d}"

        if classified == "skipped":
            # ■ 一括ダミー — 名前が偶々一致しても details には反映しない
            matched.append({
                "excel_no": r["no"], "excel_name": r["name"],
                "csv_no": no, "csv_name": match["name"],
                "no_agrees": int(r["no"]) == no,
                "status": raw_status, "classified": classified,
            })
            continue

        if classified in HAS_ANSWER_STATUSES:
            files_raw = r.get("attach_files_raw") or ""
            files_all = [s.strip() for s in re.split(r"[\r\n]+", files_raw) if s.strip()]
            files = [f for f in files_all if is_useful_attachment(f)]
            attach_dir = r["attach_dir"] or None
            # company_dir = "attachments/株式会社ナカトミ/" → "株式会社ナカトミ"
            company_dir = None
            if attach_dir:
                cd = attach_dir.replace("\\", "/").rstrip("/")
                if cd.startswith("attachments/"):
                    cd = cd[len("attachments/"):]
                company_dir = cd or None
            details[key].update({
                "has_answer": True,
                "status": classified,
                "reply_date": fmt_dt(r["reply_dt"]),
                "q1": r["q1"], "q2": r["q2"], "q3": r["q3"],
                "q4": r["q4"], "q5": r["q5"],
                "attachments": files,
                "attachment_dir": attach_dir,
                "company_dir": company_dir,
            })
        else:
            # unanswered / unknown — status だけ記録、回答系フィールドは触らない
            details[key].update({
                "has_answer": False,
                "status": classified,
            })

        matched.append({
            "excel_no": r["no"], "excel_name": r["name"],
            "csv_no": no, "csv_name": match["name"],
            "no_agrees": int(r["no"]) == no,
            "status": raw_status, "classified": classified,
        })
    return details, matched, unmatched, raw_status_counter, classified_counter, aliased


def write_csv(makers: list[dict], details: dict) -> Path:
    bak = CSV_PATH.with_suffix(".csv.bak")
    shutil.copy2(CSV_PATH, bak)
    fields = ["no", "name", "name_short", "category", "has_answer", "pamphlet_page"]
    with open(CSV_PATH, "w", encoding="utf-8", newline="") as f:
        w = csv.DictWriter(f, fieldnames=fields)
        w.writeheader()
        for m in makers:
            row = dict(m)
            key = f"{int(m['no']):03d}"
            row["has_answer"] = "true" if details[key]["has_answer"] else "false"
            w.writerow(row)
    return bak


def main():
    ap = argparse.ArgumentParser(description="Merge TWF2026 回答集約.xlsx into portal data.")
    ap.add_argument("--excel", default=str(DEFAULT_EXCEL), help="path to 回答集約.xlsx")
    ap.add_argument("--dry", action="store_true", help="preview only; don't write files")
    args = ap.parse_args()

    excel_path = Path(args.excel)
    if not excel_path.exists():
        print(f"ERROR: Excel not found: {excel_path}", file=sys.stderr)
        sys.exit(1)

    makers = load_makers()
    excel_rows = load_excel(excel_path)
    details, matched, unmatched, raw_status_counter, classified_counter, aliased = merge(makers, excel_rows)

    print(f"CSV makers loaded:  {len(makers)}  ({CSV_PATH.relative_to(ROOT)})")
    print(f"Excel rows loaded:  {len(excel_rows)}  ({excel_path})")
    print(f"Matched:            {len(matched)}  (うち alias 経由: {len(aliased)})")
    print(f"Unmatched warnings: {len(unmatched)}")
    print()
    print("--- Status breakdown (raw → classified) ---")
    for raw, n in raw_status_counter.most_common():
        cls = classify_status(raw)
        print(f"  raw={raw!r:30s} → {cls:10s}  count={n}")
    print()
    print("--- Status aggregated ---")
    for cls, n in classified_counter.most_common():
        marker = "✓" if cls in HAS_ANSWER_STATUSES else " "
        print(f"  {marker} {cls:10s}  {n}")
    print()
    print("--- Matched ---")
    print(f"{'ExNo':>4}  {'CsvNo':>5}  {'No?':>4}  {'class':>11}  Excel社名 -> CSV社名")
    for m in matched:
        flag = "ok" if m["no_agrees"] else "DIFF"
        print(f"{m['excel_no']:>4}  {m['csv_no']:>5}  {flag:>4}  {m['classified']:>11}  {m['excel_name']!s} -> {m['csv_name']!s}")

    if aliased:
        print()
        print("--- Aliased (data/maker_aliases.json 経由) ---")
        for a in aliased:
            print(f"  excel No={a['excel_no']}  classified={a['classified']}  {a['excel_name']!s} -> CSV No={a['csv_no']} {a['csv_name']!s}")

    if unmatched:
        print()
        print(f"--- Unmatched ({len(unmatched)} 件) ---")
        print("    (skipped 系 = ■ 一括 / ─ 経由 は makers.csv に対応行が無くて当然)")
        for r in unmatched:
            extra = []
            if r.get("email"):
                extra.append(f"email={r['email']}")
            if r.get("reply_email") and r.get("reply_email") != r.get("email"):
                extra.append(f"reply={r['reply_email']}")
            extra_s = "  /  " + " / ".join(extra) if extra else ""
            print(f"  excel No={r['excel_no']:>4}  classified={r['classified']:>11}  "
                  f"name={r['excel_name']!r}{extra_s}")
        # alias.json への候補ヒント (skipped 以外で名前突合できなかったもの)
        unmatched_real = [r for r in unmatched if r["classified"] != "skipped"]
        if unmatched_real:
            print()
            print(f"--- 要対応 (skipped以外で未マッチ {len(unmatched_real)} 件) ---")
            print("    名前ゆらぎなら data/maker_aliases.json に追加してください。")
            for r in unmatched_real:
                print(f"  excel No={r['excel_no']}  status={r['status']}  name={r['excel_name']!r}")

    # --- maker_overrides.json 適用 (mapper反映後の最終調整) ---
    overrides = load_overrides()
    if overrides:
        print()
        print(f"--- Overrides ({len(overrides)} 件、data/maker_overrides.json) ---")
        for key, fields in overrides.items():
            if key not in details:
                print(f"  WARN: override key {key!r} は details に存在しない (skip)")
                continue
            applied = []
            for fk, fv in fields.items():
                old = details[key].get(fk)
                details[key][fk] = fv
                if old != fv:
                    applied.append(fk)
            if applied:
                print(f"  No.{key} {details[key]['name']}: 上書き {applied}")
            else:
                print(f"  No.{key} {details[key]['name']}: (差分なし、override 不要かも)")

    answered = sum(1 for v in details.values() if v["has_answer"])
    print()
    print(f"has_answer=true (CSV側): {answered} / {len(details)}")
    print(f"  内訳: answered={sum(1 for v in details.values() if v['status'] == 'answered')} "
          f"partial={sum(1 for v in details.values() if v['status'] == 'partial')}")

    if args.dry:
        print("[DRY] No files written.")
        return

    JSON_PATH.parent.mkdir(parents=True, exist_ok=True)
    with open(JSON_PATH, "w", encoding="utf-8") as f:
        json.dump(details, f, ensure_ascii=False, indent=2)
    bak = write_csv(makers, details)

    print()
    print(f"Wrote: {JSON_PATH.relative_to(ROOT)}  ({JSON_PATH.stat().st_size:,} bytes)")
    print(f"Wrote: {CSV_PATH.relative_to(ROOT)}  (backup: {bak.relative_to(ROOT)})")


if __name__ == "__main__":
    main()
